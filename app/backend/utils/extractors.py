"""Text extractors for PDF, DOCX, TXT, XLSX and images (OCR)."""
from pathlib import Path
from typing import Tuple


def extract_pdf(path: str) -> Tuple[str, int]:
    from PyPDF2 import PdfReader
    reader = PdfReader(path)
    pages = []
    for p in reader.pages:
        try:
            pages.append(p.extract_text() or "")
        except Exception:
            pages.append("")
    return "\n\n".join(pages), len(reader.pages)


def extract_docx(path: str) -> Tuple[str, int]:
    from docx import Document as Docx
    d = Docx(path)
    text = "\n".join(para.text for para in d.paragraphs if para.text)
    return text, 1


def extract_txt(path: str) -> Tuple[str, int]:
    return Path(path).read_text(encoding="utf-8", errors="ignore"), 1


def extract_xlsx(path: str) -> Tuple[str, int]:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    parts = []
    for ws in wb.worksheets:
        parts.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            line = " | ".join("" if c is None else str(c) for c in row)
            if line.strip():
                parts.append(line)
    return "\n".join(parts), len(wb.worksheets)


def extract_image(path: str) -> Tuple[str, int]:
    import pytesseract
    from PIL import Image
    img = Image.open(path)
    return pytesseract.image_to_string(img), 1


def extract_by_type(path: str, ext: str) -> Tuple[str, int]:
    ext = ext.lower().lstrip(".")
    if ext == "pdf":
        return extract_pdf(path)
    if ext in {"docx", "doc"}:
        return extract_docx(path)
    if ext in {"xlsx", "xls"}:
        return extract_xlsx(path)
    if ext in {"png", "jpg", "jpeg", "webp", "bmp", "tif", "tiff"}:
        return extract_image(path)
    return extract_txt(path)
